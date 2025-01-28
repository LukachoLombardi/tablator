import logging
import os.path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.hyperlink import Hyperlink
from pydantic import BaseModel


def initialize_table(model_type: type(BaseModel), table_name: str = "") -> Workbook:
    wb = Workbook()
    ws = wb.active
    if table_name == "":
        ws.title = model_type.__name__
    else:
        ws.title = table_name
    title_row = ["Index"]
    dump = model_type.__fields__
    for key, value in dump.items():
        if value.title is not None:
            title_row.append(value.title)
        else:
            title_row.append(str(key))
    ws.append(title_row)
    return wb

def write_model_to_table(model: BaseModel, ws: Workbook):
    row: list = []
    # data step
    if ws.active["A1"].value == "Index":
        row.append(ws.active.max_row)
    else:
        row.append(ws.active.max_row + 1)
    for key, value in model.model_dump().items():
        if value is not None:
            row.append(str(value))
        elif "opt" in key:
            row.append("opt")
        else:
            row.append("")
    ws.active.append(row)
    # fill step
    for cell in ws.active[ws.active.max_row]:
        if isinstance(cell.value, str) and os.path.isfile(cell.value):
            cell.hyperlink = Hyperlink(ref=cell.value, target=cell.value)
            cell.font = Font(color="0000FF", underline="single")
        if cell.value is (None or ""):
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif cell.value == "opt":
            cell.value = ""
    return ws


def save_table_to_file(ws: Workbook, file_path: str):
    logging.info(f"saving a table to file {file_path}")
    ws.save(file_path)
    return file_path
